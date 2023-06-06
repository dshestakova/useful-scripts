from typing import Optional, Union
import docxtpl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

DATANAME = 'data.xlsx'
wb = load_workbook(DATANAME, data_only=True)
DOCNAME = "Template.docx"


# TODO auto redactor
def docs_redactor(DocName=DOCNAME):
    # doc = DocxTemplate(DocName)
    # paragraph.text = paragraph.text.replace('{', '{{').replace('}', '}}').replace(' ', '_')
    # doc.save('changed_Template.docx')
    ...


def sheet_function(sheet, ColumnName: str) -> list:
    return [el.value for el in sheet[ColumnName] if el.value is not None]


def get_columns_names_from_sheet(SheetName: str) -> list:
    sheet = wb[SheetName]
    # rows = sheet.iter_rows(min_row=1, max_row=get_maximum_rows(sheet))
    first_row = list(sheet.rows)[get_first_row(sheet)]
    headings = [c.value for c in first_row if c.value is not None or first_row.index(c) == 0]
    headings = [c.replace('\n', '').replace('\t', '') if c is not None else c for c in headings]
    return headings


def get_sheets_names_from_table():
    return wb.sheetnames


def get_names() -> Optional[list]:
    SheetNames = get_sheets_names_from_table()
    for sheetname in SheetNames:
        ColumnsNames = get_columns_names_from_sheet(sheetname)
        name_column = list(filter(lambda x: "фамилия" or "ФИО" in x.lower(), ColumnsNames))
        if len(name_column) != 0:
            col_ind = ColumnsNames.index(name_column[0])
            return sheet_function(wb[sheetname], get_column_letter(col_ind + 1))
    return None


def get_first_row(sheet):
    rows = 0
    for row in sheet:
        if sum([col.value is not None for col in row]) > 1:
            return rows
        rows += 1


def get_maximum_rows(sheet):
    flag = 0
    for max_row, row in enumerate(sheet, 1):
        if not all(col.value is None for col in row):
            flag += 1 if flag != 2 else 0
        elif all(col.value is None for col in row) and flag == 2:
            return max_row


def find_by_value(sheetname: str, name: str) -> Union[Union[int, any], Union[None, None]]:
    sheet = wb[sheetname]
    rows = get_maximum_rows(sheet)
    first_row = get_first_row(sheet)
    columns = get_columns_names_from_sheet(sheetname)
    for row in range(first_row, rows):
        for col in range(len(columns)):
            var = sheet.cell(row + 1, col + 1).value
            if var is None:
                continue
            var = var.replace('\t', '').replace('\n', '')
            if var == name:
                return row, get_column_letter(col + 1)
    return None, None


def create_dictionary(names) -> Optional[dict]:
    context = {}
    SheetsNames = get_sheets_names_from_table()
    all_names = get_names()[1:]
    for name in names:
        context[name] = {}
        if name in all_names:
            for sheetname in SheetsNames[:1]:
                context[name][sheetname] = {}
                row, column = find_by_value(sheetname, name)

                columns = get_columns_names_from_sheet(sheetname)
                columns.remove(wb[sheetname][column][get_first_row(wb[sheetname])].value)
                for col in columns:
                    if col is None:
                        continue
                    _, column_ind = find_by_value(sheetname, col)
                    val = wb[sheetname][column_ind][row].value
                    val = val.replace('\n', '').replace('\t', '').replace('#N/A',
                                                                          'не заполнено') if val is not None else 'не заполнено'
                    col = col.strip().replace(' ', '_').replace('?', '').replace(',', '_').replace('(-а)', '').replace(
                        '/', '_')
                    context[name][sheetname][col] = val

                context[name][sheetname]['ФИ'] = name
    return context


if __name__ == '__main__':

    names = ['Карлюк Ольга']
    dict_ = create_dictionary(names)
    print(dict_)
    for name in dict_:
        for sheet in dict_[name]:
            context = dict_[name][sheet]

            doc = docxtpl.DocxTemplate("template.docx")
            doc.render(context)
            doc.save(f"Индивидуальный_отчет_{name}.docx")
